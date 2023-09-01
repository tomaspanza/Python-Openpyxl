import openpyxl
from openpyxl.styles import Font
import os


# Main menu
def mostrar_menu():
    print("\nMenú:")
    print("1. Agregar Empleados")
    print("2. Modificar Empleados")
    print("3. Borrar Empleados")
    print("4. Salir")


# Guardado de cambios en el excel
def guardar_archivo():
    workbook.save("Empleados.xlsx")
    print("Cambios guardados en 'Empleados.xlsx'.")


# Solicitar al usuario que ingrese los datos
def agregar_empleado():
    numero_legajo = input("Ingrese el número de legajo: ")
    nombre = input("Ingrese el nombre: ")
    edad = int(input("Ingrese la edad: "))
    area = input("Ingrese el área: ")
    horas_diarias = float(input("Ingrese las horas diarias: "))
    pago_por_hora = float(input("Ingrese el pago por hora: "))
    dias_trabajados = int(input("Ingrese los días trabajados: "))
    sueldo_mensual = horas_diarias * pago_por_hora * dias_trabajados

    data = [
        numero_legajo,
        nombre,
        edad,
        area,
        horas_diarias,
        pago_por_hora,
        dias_trabajados,
        sueldo_mensual,
    ]
    sheet.append(data)
    guardar_archivo()
    print("Empleado se agrego con éxito!!")


# Modificar empleado


def modificar_empleado():
    legajo_modificar = input("Ingrese el número de legajo del empleado a modificar: ")
    encontrado = False

    for row_index, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if row[0] == legajo_modificar:
            encontrado = True
            print(f"Empleado encontrado: {row}")
            nuevo_nombre = input("Ingrese el nuevo nombre: ")
            nuevo_edad = int(input("Ingrese la nueva edad: "))
            nuevo_area = input("Ingrese la nueva área: ")
            nueva_horas_diarias = float(input("Ingrese las nuevas horas diarias: "))
            nuevo_pago_por_hora = float(input("Ingrese el nuevo pago por hora: "))
            nuevos_dias_trabajados = int(input("Ingrese los nuevos días trabajados: "))
            nuevo_sueldo_mensual = (
                nueva_horas_diarias * nuevo_pago_por_hora * nuevos_dias_trabajados
            )

            # Crear una nueva lista con los valores actualizados
            nueva_fila = [
                row[0],
                nuevo_nombre,
                nuevo_edad,
                nuevo_area,
                nueva_horas_diarias,
                nuevo_pago_por_hora,
                nuevos_dias_trabajados,
                nuevo_sueldo_mensual,
            ]

            # Reemplazar la fila en la hoja con los nuevos valores
            for col_index, value in enumerate(nueva_fila, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)

            guardar_archivo()
            print("Empleado modificado con éxito.")
            break

    if not encontrado:
        print("Empleado no encontrado.")


# Borrar empleado


def borrar_empleado():
    legajo_borrar = input("Ingrese el número de legajo del empleado a borrar: ")
    encontrado = False

    for row_index, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if row[0] == legajo_borrar:
            encontrado = True
            print(f"Empleado a borrar: {row}")
            confirmar = input(
                "¿Está seguro de que desea borrar este empleado? (S/N): "
            ).lower()
            if confirmar == "s":
                sheet.delete_rows(row_index)
                guardar_archivo()
                print("Empleado borrado con éxito.")
            else:
                print("Borrado cancelado.")
            break

    if not encontrado:
        print("Empleado no encontrado.")


# Verificar si el archivo Excel ya existe
if os.path.exists("Empleados.xlsx"):
    workbook = openpyxl.load_workbook("Empleados.xlsx")
    sheet = workbook.active
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Empleados_Mañana"
    headers = [
        "Numero de Legajo",
        "Nombre",
        "Edad",
        "Area",
        "Horas diarias",
        "Pago por hora",
        "Dias trabajados",
        "Sueldo mensual",
    ]
    sheet.append(headers)
    guardar_archivo()

mostrar_menu()


# Aplicar formato a los encabezados en negrita
for cell in sheet[1]:
    cell.font = Font(bold=True)


while True:
    opcion = input("\nIngrese la opción deseada: ")

    if opcion == "1":
        agregar_empleado()
    elif opcion == "2":
        modificar_empleado()
    elif opcion == "3":
        borrar_empleado()
    elif opcion == "4":
        print("Muchas gracias por usar el sistema!")
        break
    else:
        print("Opción no válida. Por favor, elija una opción del menú.")
